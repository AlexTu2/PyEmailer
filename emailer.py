import sys
import os
import string
import re
import argparse
import tkinter as tk
from tkinter import filedialog
from contextlib import suppress
import openpyxl
import ezgmail
import bs4

class PatchError(Exception):
    """Custom exception raised when there is an issue applying a patch."""
    pass


def patch():
    """
    Apply a patch to the ezgmail library by modifying its version information.

    This function reads the content of the ezgmail library, checks its version,
    and applies a patch if the version is compatible. The patched version is then saved.

    Raises:
        PatchError: If the patch cannot be applied due to an invalid version.
    """
    site_packages = sys.path[5]
    ezgmail_code_path = os.path.join(site_packages, "ezgmail", "__init__.py")
    # __version__ = \"2022.10.10\"
    with open(ezgmail_code_path, "r") as f:
        code = f.read()
    if '__version__ = "2022.10.10"' in code:
        pass
    elif '__version__ = "2022.10.10.PATCHED"' in code:
        return
    else:
        raise PatchError("Unable to apply patch! Invalid version!")
    lines = code.split("\n")
    lines[6] = '__version__ = "2022.10.10.PATCHED"'
    lines[491] = "#" + lines[491]
    lines[525] = "#" + lines[525]
    code_new = "\n".join(lines)
    with open(ezgmail_code_path, "w") as f:
        f.truncate()
        f.write(code_new)


def prettify_except(soup_obj: bs4.BeautifulSoup, tag_name: str) -> str:
    """
    Prettify the BeautifulSoup object, excluding a specific tag.

    Args:
        soup_obj (bs4.BeautifulSoup): The BeautifulSoup object.
        tag_name (str): The tag to exclude.

    Returns:
        str: The prettified HTML text excluding the specified tag.
    """
    # https://stackoverflow.com/a/69589000/9091833
    regex_string = "<{0}>.*<\/{0}>".format(tag_name)
    regex = re.compile(regex_string, re.DOTALL)
    replacing_txt = str(getattr(soup_obj, tag_name))
    return re.sub(regex, replacing_txt, soup_obj.prettify())


# batch of 20, complete list, range of rows, start row then quant,
# add sent email date & time
# use sent date to confirm if email needs to be sent again


def mail_from_excel(mail_list, template, sig, _closing, _name):
    """
    Compose and draft emails based on data from an Excel sheet.

    Args:
        mail_list (str): Path to the mailing list Excel sheet.
        template (str): Path to the email template file.
        sig (str): Path to the signature file.
        _closing (str): The closing phrase for the email.
        _name (str): The name to come after the closing.

    Notes:
        The email content is composed using the provided template, signature, closing,
        and name. Draft emails are created for each row in the Excel sheet.

    """
    wb = openpyxl.load_workbook(mail_list)
    try:
        sheet = wb["Sheet1"]
        data = tuple(sheet.rows)
    finally:
        wb.close()

    with open(template, "r") as infile:
        template = infile.read()
        soup = bs4.BeautifulSoup(template, "html.parser")

    with open(sig, "r") as infile:
        sig = infile.read()

    soup.body.append(bs4.BeautifulSoup(sig, "html.parser"))
    with open(r"Templates\message.html", "w") as outfile:
        outfile.write(prettify_except(soup, "body"))
    with open(r"Templates\message.html", "r") as infile:
        message = infile.read()

    field_names = [v[1] for v in string.Formatter().parse(template)]

    print(f"Rows in sheet {sheet.max_row}")
    # Start at 1 to skip header, "name" and "email"
    for i in range(1, sheet.max_row):
        ezgmail.draft(
            data[i][1].value,
            f"Mail for {data[i][0].value}",
            message.format(closing=_closing, name=_name),
            mimeSubtype="html",
        )


def logout():
    """
    Log out the user by removing the token.json file.

    Notes:
        If the token.json file is present, it will be removed.
    """
    with suppress(OSError):
        os.remove("token.json")


def user_auth():
     """
    Authenticate the user with the ezgmail library.

    This function initializes the ezgmail library and prompts the user to log out
    or continue with the email program.
    """
    while True:
        ezgmail.init()
        print("=" * 80)
        print(
            f"Python Emailer, you're currently logged in as: {ezgmail.EMAIL_ADDRESS}"
        )
        print("=" * 80 + "\n\n")

        # prompt for user change
        while True:
            logout_choice = input("Do you want to logout? ((y)es / (N)o): ").lower()
            if logout_choice in ("yes", "no", "y", "n", ""):
                break
            print("invalid input")

        if logout_choice in ("yes", "y"):
            logout()
            continue
        else:
            print("Continuing with email program... \n")
            break


def prompt_for_file(prompt):
    """
    Prompt the user to enter a file path.

    Args:
        msg (str): The prompt message.

    Returns:
        str: The entered file path.
    """
    while True:
        file_path = input(f"\n{prompt}: ")
        if os.path.isfile(file_path):
            return file_path
        else:
            print("Invalid file path. Please enter a valid file path.")


def select_file_dialog(prompt, root, cwd):
    print(prompt)
    return filedialog.askopenfilename(parent=root, initialdir=cwd)


def main():
    # patch()

    user_auth()
    cwd = os.getcwd()
    print(f"The cwd is: {cwd}")

    parser = argparse.ArgumentParser()
    parser.add_argument(
        "-m",
        "--mail_list",
        help="The mailing list excel sheet (.xlsx,.xlsm,.xltx,.xltm)",
    )
    args = parser.parse_args()
    # args = parser.parse_args([r"Mail lists\example.xlsx"])

    # Determine if user will use gui tk file prompt or cmdline input
    while True:
        gui_choice = input(
            "Do you wish to use the GUI File picker? (Y)es/(n)o: "
        ).lower()
        if gui_choice in ("yes", "no", "y", "n", ""):
            break
        print("Invalid choice, please retry. ")

    if gui_choice in ("yes", "y", ""):
        using_gui_filedialog = True
        root = tk.Tk()
        root.withdraw()
        root.lift()
        root.focus_force()
    else:
        using_gui_filedialog = False
    print("=" * 80 + "\n")

    # get/prompt for mail list excel sheet (.xlsx,.xlsm,.xltx,.xltm)
    if args.mail_list:
        # print("Sys.argv found!")
        mail_list = args.mail_list
    else:
        if using_gui_filedialog:
            mail_list = select_file_dialog("Select a mailing list file", root, cwd)
        else:
            mail_list = prompt_for_file("Enter mailing list file")
    print(f"Mailing list file: {mail_list}\n")

    # prompt for template
    if using_gui_filedialog:
        template = select_file_dialog("Select a template file", root, cwd)
    else:
        template = prompt_for_file("Enter template file")
    print(f"Template file: {template}\n")

    # prompt for sig name
    if using_gui_filedialog:
        sig = select_file_dialog("Select a signature file", root, cwd)
    else:
        sig = prompt_for_file("Enter signature file")
    print(f"Signature file: {sig}\n")

    # Prompt for closing
    print("=" * 80 + "\n")
    closing = input("Enter a closing (Best, Best regards, Signed, etc.): ")

    # Prompt for name
    name = input("Enter a name to come after the closing: ")

    print("\n\n")
    mail_from_excel(mail_list, template, sig, closing, name)

    input("Enter to exit: ")


if __name__ == "__main__":
    os.chdir(os.path.dirname(os.path.abspath(__file__)))
    main()
